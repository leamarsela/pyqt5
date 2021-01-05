import sys
import os
from PyQt5 import Qt
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from math import sqrt, pi
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator, AutoMinorLocator
from matplotlib.offsetbox import AnchoredText, TextArea, DrawingArea, OffsetImage, AnnotationBbox
import matplotlib.image as mpimg
from scipy.stats import linregress
from math import asin, degrees, cos, radians
from fpdf import FPDF
from PyPDF2 import PdfFileMerger, PdfFileReader, PdfFileWriter
import numpy as np



class MainForm(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon('nglogo1.ico'))
        self.setupGui()

    def setupGui(self):
        self.resize(600, 500)
        self.setWindowTitle('NGplot - Main Window')
        self.setCenter()

        self.startext1 = QLabel(self)
        self.startext1.setText('NGplot')
        self.startext1.setFont(QFont('Times', 36, QFont.Bold))
        self.startext1.setAlignment(Qt.AlignCenter)

        self.startext2 = QLabel(self)
        self.startext2.setText('Triaxial Consolidated Undrained Plot')
        self.startext2.setFont(QFont('Times', 10, QFont.Bold))
        self.startext2.setAlignment(Qt.AlignCenter)

        self.startbutton = QPushButton('Start')
        self.startbutton.setToolTip('Start')
        self.startbutton.clicked.connect(self.startbuttonClick)

        layout = QVBoxLayout(self)
        layout.addWidget(self.startext1)
        layout.addWidget(self.startext2)
        layout.addWidget(self.startbutton)

        self.setLayout(layout)

    def startbuttonClick(self):
        self.inputform = InputForm()
        self.inputform.show()

    def setCenter(self):
        desktop = QDesktopWidget()
        screenwidth = desktop.screen().width()
        screenheight = desktop.screen().height()

        self.setGeometry(int((screenwidth - self.width()) / 2), int((screenheight - self.height()) / 2),
                        (self.width()),
                        (self.height()))

sqrti = lambda ti: sqrt(ti)
class InputForm(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon('nglogo1.ico'))
        self.setupGuiInput()

    def setupGuiInput(self):
        self.resize(600, 500)
        self.setWindowTitle('NGplot - Input Window')
        self.setInputCenter()

        self.openfilebutton = QPushButton('Open File TX-CU')
        self.openfilebutton.setToolTip('Open File')
        # self.openfilebutton.setIcon(QIcon('open.png'))
        self.openfilebutton.clicked.connect(self.openfilebuttonClick)

        self.openfiletext = QLineEdit(self)
        self.openfiletext.setDisabled(True)

        self.labellistprintout = QLabel('Print Out:')

        self.listprintout = QListWidget()
        self.listprintout.setDisabled(True)
        self.listprintout.verticalScrollBar().setValue(self.listprintout.verticalScrollBar().maximum())

        self.progressbar = QProgressBar()
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(10000)
        self.progressbar.setValue(0)

        self.runfilebutton = QPushButton('Run')
        self.runfilebutton.setToolTip('Run file')
        # self.runfilebutton.setIcon(QIcon('run.png'))
        self.runfilebutton.setDisabled(True)
        self.runfilebutton.clicked.connect(self.runfilebuttonClick)

        self.showresultbutton = QPushButton('Show Result')
        self.showresultbutton.setToolTip('Show Result')
        self.showresultbutton.setDisabled(True)
        self.showresultbutton.clicked.connect(self.showresultbuttonClick)

        layoutInput = QVBoxLayout()
        layoutInput.addWidget(self.openfilebutton)
        layoutInput.addWidget(self.openfiletext)
        layoutInput.addWidget(self.labellistprintout)
        layoutInput.addWidget(self.listprintout)

        hbox = QHBoxLayout()
        hbox.addWidget(self.progressbar)
        hbox.addStretch()
        hbox.addWidget(self.runfilebutton)
        hbox.addWidget(self.showresultbutton)

        alllayout = QVBoxLayout()
        alllayout.addLayout(layoutInput)
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        alllayout.addWidget(line)
        alllayout.addLayout(hbox)

        self.setLayout(alllayout)

    def openfilebuttonClick(self):
        import os
        filename = QFileDialog.getOpenFileName(self, 'Open File', os.curdir, '(*.xlsx *.xls)')
        if not filename[0]:
            return
        else:
            self.openfiletext.setText(str(filename[0]))
            self.openfiletext.setStyleSheet("background-color: White")
            self.openfiletext.setDisabled(True)
            self.activeopenfiletext()
        
        global pathfile

        pathfile = filename[0]

    def activeopenfiletext(self):
        if len(self.openfiletext.text().strip()) != 0:
            self.runfilebutton.setDisabled(False)
        else:
            self.runfilebutton.setDisabled(True)

        self.listprintout.setDisabled(True)
        self.listprintout.clear()
        self.progressbar.setValue(0)

    def runfilebuttonClick(self):
        try:
            self.calcBvalue()
            self.stage1()
            self.stage2()
            self.stage3()
            self.mohrlinetotal()
            self.mohrlineeffective()
            self.plotfigure()
            self.printoutput()
            self.printrawdata()    
            self.mergerawfig()
            self.listprintout.setDisabled(False)
            self.showresultbutton.setDisabled(False)
        except:
            QMessageBox.warning(self, 'Warning', 'Error - Please check the input !!!')
            self.listprintout.setDisabled(True)
            self.showresultbutton.setDisabled(True)

    def calcBvalue(self):
        self.calcfilename = ((str(self.openfiletext.text())).split('/'))[
            len(list((str(self.openfiletext.text())).split('/'))) - 1]
        
        self.wbfile = load_workbook(self.calcfilename, data_only=True)

        # get data from header
        self.sampleid = self.wbfile[self.wbfile.sheetnames[0]].cell(1, 2).value
        self.depthfrom = self.wbfile[self.wbfile.sheetnames[0]].cell(3, 2).value
        self.depthto = self.wbfile[self.wbfile.sheetnames[0]].cell(3, 3).value
        self.calibration = self.wbfile[self.wbfile.sheetnames[0]].cell(4, 2).value      # unit in kg/div
        self.calibrationstrain = 0.01           # unit in mm/div
        self.diametersample = self.wbfile[self.wbfile.sheetnames[0]].cell(6, 2).value * 10
        self.heightsample = self.wbfile[self.wbfile.sheetnames[0]].cell(6, 3).value * 10

        #global parameter to submitted in Header 
        global idsample
        global fromdepth
        global todepth

        idsample = self.sampleid
        fromdepth = self.depthfrom
        todepth = self.depthto
        ################ PART OF SATURATION ################
        # get data from saturation
        self.maxrowsaturation = self.wbfile[self.wbfile.sheetnames[1]].max_row

        self.cpvalue = []
        self.pwpvalue = []
        for i in range(1, self.maxrowsaturation):
            self.cpvalue.append(self.wbfile[self.wbfile.sheetnames[1]].cell(1 + i, 1).value)
            self.pwpvalue.append(self.wbfile[self.wbfile.sheetnames[1]].cell(1 + i, 3).value)
        
        self.deltapwpvalue = []
        for i in range(self.maxrowsaturation - 2):
            if i == 0:
                self.deltapwpvalue.append(0)
            self.deltapwpvalue.append(self.pwpvalue[i + 1] - self.pwpvalue[i])

        self.deltacp = []
        for i in range(self.maxrowsaturation - 1):
            try:
                if i == 0:
                    self.deltacp.append(0)
                elif i == 1:
                    self.deltacp.append(self.cpvalue[i] - 0)
                elif i > 0:
                    self.deltacp.append(self.cpvalue[i] - self.cpvalue[i - 2])
            except:
                self.deltacp.append(0)

        self.sortcpvalue = []
        self.bvalue = []
        for i in range(int((self.maxrowsaturation - 1) / 2)):
            try:
                self.bvalue.append(self.deltapwpvalue[2 * i + 1] / self.deltacp[2 * i + 1])
                self.sortcpvalue.append(self.cpvalue[2 * i + 1])
            except:
                self.bvalue.append(0)
                self.sortcpvalue.append(0)
        
    def stage1(self):
        rowconsol1 = 0
        min_row = 4
        max_row = 25
        min_col = 1
        for i in self.wbfile[self.wbfile.sheetnames[2]].iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
                                                                  max_col=min_col):
            if self.wbfile[self.wbfile.sheetnames[2]].cell(1+rowconsol1, min_col).value != None:
                rowconsol1 += 1
            rowconsol1 += 0

        self.maxrowconsol1 = rowconsol1

        self.time1 = []
        self.squaretime1 = []
        self.pwpconsol1 = []
        self.volchangeconsol1 = []
        for i in range(5, self.maxrowconsol1 + 1):
            self.time1.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 1).value)
            self.squaretime1.append(sqrti(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 1).value))
            self.pwpconsol1.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 2).value)
            self.volchangeconsol1.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 3).value)

        self.deltavolchange1 = abs(self.volchangeconsol1[len(self.volchangeconsol1)-1] - self.volchangeconsol1[0])
        self.deltapwpconsol1 = abs(self.pwpconsol1[len(self.volchangeconsol1)-1] - self.pwpconsol1[0])

        self.deltavolchangeconsol1 = []
        self.dissipation1 = []
        for i in range(len(self.time1)):
            self.deltavolchangeconsol1.append(abs(self.volchangeconsol1[i] - self.volchangeconsol1[0]))
            self.dissipation1.append((self.pwpconsol1[0] - self.pwpconsol1[i]) * 100 / self.deltapwpconsol1)

        self.cp1 = self.wbfile[self.wbfile.sheetnames[2]].cell(2, 2).value * 100.
        self.bp1 = self.wbfile[self.wbfile.sheetnames[2]].cell(3, 2).value * 100.
        self.ep1 = self.cp1 - self.bp1

        rowshear1 = 0
        min_row = 2
        max_row = 300
        min_col = 1
        for i in self.wbfile[self.wbfile.sheetnames[3]].iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
                                                                  max_col=min_col):
            if self.wbfile[self.wbfile.sheetnames[3]].cell(1 + rowshear1, min_col).value != None:
                rowshear1 += 1
            rowshear1 += 0

        self.maxrowshear1 = rowshear1

        self.straingauge1 = []
        self.dialload1 = []
        self.pwpshear1 = []
        for i in range(3, self.maxrowshear1 + 1):
            self.straingauge1.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 1).value)
            self.dialload1.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 2).value)
            self.pwpshear1.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 3).value)

        self.deltaheight1 = []
        self.deltaepsilon1 = []
        self.areacf1 = []
        for i in range(len(self.straingauge1)):
            self.deltaheight1.append((self.straingauge1[i] - self.straingauge1[0]) * self.calibrationstrain)
            self.deltaepsilon1.append(self.deltaheight1[i] / self.heightsample)
            if i == 0:
                self.areacf1.append(1)
            else:
                self.areacf1.append(self.areacf1[0] - self.deltaepsilon1[i])

        self.area01 = 0.25 * pi * pow(self.diametersample, 2)
        self.volume01 = self.area01 * self.heightsample / 1000
        self.epsilonV1 = self.deltavolchange1 * 100 / self.volume01
        self.lengthCc1 = self.heightsample * (1 - (self.epsilonV1 / 300))
        self.areaCc1 = self.area01 * (1 - (2 * self.epsilonV1 / 300))
        self.lengthCs1 = self.lengthCc1 - self.deltaheight1[len(self.deltaheight1) - 1]

        self.areacorected1 = []
        for i in range(len(self.straingauge1)):
            if i == 0:
                self.areacorected1.append(self.areaCc1)
            else:
                self.areacorected1.append(self.areacorected1[0] / self.areacf1[i])

        self.diameter01 = sqrt(self.areacorected1[len(self.straingauge1) - 1] / (0.25 * pi))

        self.load1 = []
        for i in range(len(self.straingauge1)):
            self.load1.append(self.dialload1[i] * self.calibration * 10)              #unit in Newton

        self.deviatorstress1 = []
        for i in range(len(self.straingauge1)):
            self.deviatorstress1.append(self.load1[i] * 1000 / self.areacorected1[i])     #unit in kPa

        self.strainvalue1 = []
        for i in range(len(self.straingauge1)):
            self.strainvalue1.append(self.deltaheight1[i] * 100 / self.heightsample)

        self.deltapwpshear1 = []
        for i in range(len(self.straingauge1)):
            self.deltapwpshear1.append((self.pwpshear1[i] - self.pwpshear1[0]) * 100)   #unit in kPa

        self.sigma1T1 = []
        for i in range(len(self.straingauge1)):
            self.sigma1T1.append(self.ep1 + self.deviatorstress1[i])

        self.sigma1E1 = []
        for i in range(len(self.straingauge1)):
            self.sigma1E1.append(self.sigma1T1[i] - self.deltapwpshear1[i])

        self.sigma3E1 = []
        for i in range(len(self.straingauge1)):
            self.sigma3E1.append(self.sigma1E1[i] - self.deviatorstress1[i])

    def stage2(self):
        rowconsol2 = 0
        min_row = 4
        max_row = 25
        min_col = 5
        for i in self.wbfile[self.wbfile.sheetnames[2]].iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
                                                                  max_col=min_col):
            if self.wbfile[self.wbfile.sheetnames[2]].cell(1 + rowconsol2, min_col).value != None:
                rowconsol2 += 1
            rowconsol2 += 0

        self.maxrowconsol2 = rowconsol2

        self.time2 = []
        self.squaretime2 = []
        self.pwpconsol2 = []
        self.volchangeconsol2 =[]
        for i in range(5, self.maxrowconsol2 + 1):
            self.time2.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 5).value)
            self.squaretime2.append(sqrti(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 5).value))
            self.pwpconsol2.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 6).value)
            self.volchangeconsol2.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 7).value)

        self.deltavolchange2 = abs(self.volchangeconsol2[len(self.volchangeconsol2) - 1] - self.volchangeconsol2[0])
        self.deltapwpconsol2 = abs(self.pwpconsol2[len(self.volchangeconsol2) - 1] - self.pwpconsol2[0])

        self.deltavolchangeconsol2 = []
        self.dissipation2 = []
        for i in range(len(self.time2)):
            self.deltavolchangeconsol2.append(abs(self.volchangeconsol2[i] - self.volchangeconsol2[0]))
            self.dissipation2.append((self.pwpconsol2[0] - self.pwpconsol2[i]) * 100 / self.deltapwpconsol2)

        self.cp2 = self.wbfile[self.wbfile.sheetnames[2]].cell(2, 6).value * 100.
        self.bp2 = self.wbfile[self.wbfile.sheetnames[2]].cell(3, 6).value * 100.
        self.ep2 = self.cp2 - self.bp2

        rowshear2 = 0
        min_row = 2
        max_row = 300
        min_col = 5
        for i in self.wbfile[self.wbfile.sheetnames[3]].iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
                                                                  max_col=min_col):
            if self.wbfile[self.wbfile.sheetnames[3]].cell(1 + rowshear2, min_col).value != None:
                rowshear2 += 1
            rowshear2 += 0

        self.maxrowshear2 = rowshear2

        self.straingauge2 = []
        self.dialload2 = []
        self.pwpshear2 = []
        for i in range(3, self.maxrowshear2 + 1):
            self.straingauge2.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 5).value)
            self.dialload2.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 6).value)
            self.pwpshear2.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 7).value)

        self.deltaheight2 = []
        self.deltaepsilon2 = []
        self.areacf2 = []
        for i in range(len(self.straingauge2)):
            self.deltaheight2.append(self.straingauge2[i] * self.calibrationstrain)
            self.deltaepsilon2.append(self.deltaheight2[i] / self.lengthCs1)
            if i == 0:
                self.areacf2.append(1)
            else:
                self.areacf2.append(self.areacf2[0] - self.deltaepsilon2[i])

        self.heightsample02 = self.lengthCs1
        self.diametersample2 = self.diameter01
        self.area02 = self.areacorected1[len(self.straingauge1) - 1]
        self.volume02 = self.area02 * self.lengthCs1 / 1000
        self.epsilonV2 = self.deltavolchange2 * 100 / self.volume02
        self.lengthCc2 = self.heightsample02 * (1 - (self.epsilonV2 / 300))
        self.areaCc2 = self.area02 * (1 - (2 * self.epsilonV2 / 300))
        self.lengthCs2 = self.lengthCc2 - (self.deltaheight2[len(self.deltaheight2) - 1] - self.deltaheight2[0])

        self.areacorected2 = []
        for i in range(len(self.straingauge2)):
            if i == 0:
                self.areacorected2.append(self.areaCc2)
            else:
                self.areacorected2.append(self.areacorected2[0] / self.areacf2[i])

        self.diameter02 = sqrt(self.areacorected2[len(self.straingauge2) - 1] / (0.25 * pi))

        self.load2 = []
        for i in range(len(self.straingauge2)):
            self.load2.append(self.dialload2[i] * self.calibration * 10)            #unit in Newton

        self.deviatorstress2 = []
        for i in range(len(self.straingauge2)):
            self.deviatorstress2.append(self.load2[i] * 1000 / self.areacorected2[i])   #unit in kPa

        self.strainvalue2 = []
        for i in range(len(self.straingauge2)):
            self.strainvalue2.append(self.deltaheight2[i] * 100 / self.heightsample02)

        self.deltapwpshear2 = []
        for i in range(len(self.straingauge2)):
            self.deltapwpshear2.append((self.pwpshear2[i] - self.pwpshear2[0]) * 100)   #unit in kPa

        self.sigma1T2 = []
        for i in range(len(self.straingauge2)):
            self.sigma1T2.append(self.ep2 + self.deviatorstress2[i])

        self.sigma1E2 = []
        for i in range(len(self.straingauge2)):
            self.sigma1E2.append(self.sigma1T2[i] - self.deltapwpshear2[i])

        self.sigma3E2 = []
        for i in range(len(self.straingauge2)):
            self.sigma3E2.append(self.sigma1E2[i] - self.deviatorstress2[i])

    def stage3(self):
        rowconsol3 = 0
        min_row = 4
        max_row = 25
        min_col = 9
        for i in self.wbfile[self.wbfile.sheetnames[2]].iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
                                                                  max_col=min_col):
            if self.wbfile[self.wbfile.sheetnames[2]].cell(1 + rowconsol3, min_col).value != None:
                rowconsol3 += 1
            rowconsol3 += 0

        self.maxrowconsol3 = rowconsol3

        self.time3 = []
        self.squaretime3 = []
        self.pwpconsol3 = []
        self.volchangeconsol3 =[]
        for i in range(5, self.maxrowconsol3 + 1):
            self.time3.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 9).value)
            self.squaretime3.append(sqrti(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 9).value))
            self.pwpconsol3.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 10).value)
            self.volchangeconsol3.append(self.wbfile[self.wbfile.sheetnames[2]].cell(i, 11).value)

        self.deltavolchange3 = abs(self.volchangeconsol3[len(self.volchangeconsol3) - 1] - self.volchangeconsol3[0])
        self.deltapwpconsol3 = abs(self.pwpconsol3[len(self.volchangeconsol3) - 1] - self.pwpconsol3[0])

        self.deltavolchangeconsol3 = []
        self.dissipation3 = []
        for i in range(len(self.time3)):
            self.deltavolchangeconsol3.append(abs(self.volchangeconsol3[i] - self.volchangeconsol3[0]))
            self.dissipation3.append((self.pwpconsol3[0] - self.pwpconsol3[i]) * 100 / self.deltapwpconsol3)

        self.cp3 = self.wbfile[self.wbfile.sheetnames[2]].cell(2, 10).value * 100.
        self.bp3 = self.wbfile[self.wbfile.sheetnames[2]].cell(3, 10).value * 100.
        self.ep3 = self.cp3 - self.bp3

        rowshear3 = 0
        min_row = 2
        max_row = 300
        min_col = 9
        for i in self.wbfile[self.wbfile.sheetnames[3]].iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
                                                                  max_col=min_col):
            if self.wbfile[self.wbfile.sheetnames[3]].cell(1 + rowshear3, min_col).value != None:
                rowshear3 += 1
            rowshear3 += 0

        self.maxrowshear3 = rowshear3

        self.straingauge3 = []
        self.dialload3 = []
        self.pwpshear3 = []
        for i in range(3, self.maxrowshear3 + 1):
            self.straingauge3.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 9).value)
            self.dialload3.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 10).value)
            self.pwpshear3.append(self.wbfile[self.wbfile.sheetnames[3]].cell(i, 11).value)

        self.deltaheight3 = []
        self.deltaepsilon3 = []
        self.areacf3 = []
        for i in range(len(self.straingauge3)):
            self.deltaheight3.append(self.straingauge3[i] * self.calibrationstrain)
            self.deltaepsilon3.append(self.deltaheight3[i] / self.lengthCs2)
            if i == 0:
                self.areacf3.append(1)
            else:
                self.areacf3.append(self.areacf3[0] - self.deltaepsilon3[i])

        self.heightsample03 = self.lengthCs2
        self.diametersample3 = self.diameter02
        self.area03 = self.areacorected2[len(self.straingauge2) - 1]
        self.volume03 = self.area03 * self.lengthCs2 / 1000
        self.epsilonV3 = self.deltavolchange3 * 100 / self.volume03
        self.lengthCc3 = self.heightsample03 * (1 - (self.epsilonV3 / 300))
        self.areaCc3 = self.area03 * (1 - (2 * self.epsilonV3 /300))
        self.lengthCs3 = self.lengthCc3 - (self.deltaheight3[len(self.deltaheight3) - 1] - self.deltaheight3[0])

        self.areacorected3 = []
        for i in range(len(self.straingauge3)):
            if i == 0:
                self.areacorected3.append(self.areaCc3)
            else:
                self.areacorected3.append(self.areacorected3[0] / self.areacf3[i])

        self.diameter03 = sqrt(self.areacorected3[len(self.straingauge3) - 1] / (0.25 * pi))

        self.load3 = []
        for i in range(len(self.straingauge3)):
            self.load3.append(self.dialload3[i] * self.calibration * 10)          #unit in Newton

        self.deviatorstress3 = []
        for i in range(len(self.straingauge3)):
            self.deviatorstress3.append(self.load3[i] * 1000 / self.areacorected3[i])   #unit in kPa

        self.strainvalue3 = []
        for i in range(len(self.straingauge3)):
            self.strainvalue3.append(self.deltaheight3[i] * 100 / self.heightsample03)

        self.deltapwpshear3 = []
        for i in range(len(self.straingauge3)):
            self.deltapwpshear3.append((self.pwpshear3[i] - self.pwpshear3[0]) * 100)   #unit in kPa

        self.sigma1T3 = []
        for i in range(len(self.straingauge3)):
            self.sigma1T3.append(self.ep3 + self.deviatorstress3[i])

        self.sigma1E3 = []
        for i in range(len(self.straingauge3)):
            self.sigma1E3.append(self.sigma1T3[i] - self.deltapwpshear3[i])

        self.sigma3E3 = []
        for i in range(len(self.straingauge3)):
            self.sigma3E3.append(self.sigma1E3[i] - self.deviatorstress3[i])

    def mohrlinetotal(self):
        self.mohrsigma1T1 = max(self.deviatorstress1) + self.ep1
        self.mohrsigma1T2 = max(self.deviatorstress2) + self.ep2
        self.mohrsigma1T3 = max(self.deviatorstress3) + self.ep3

        self.constaT1 = self.ep1 + (self.mohrsigma1T1 - self.ep1) / 2
        self.constaT2 = self.ep2 + (self.mohrsigma1T2 - self.ep2) / 2
        self.constaT2 = self.ep3 + (self.mohrsigma1T3 - self.ep3) / 2

        self.constrT1 = (self.mohrsigma1T1 - self.ep1) / 2
        self.constrT2 = (self.mohrsigma1T2 - self.ep2) / 2
        self.constrT3 = (self.mohrsigma1T3 - self.ep3) / 2

        self.constpT1 = (self.mohrsigma1T1 + self.ep1) / 2
        self.constpT2 = (self.mohrsigma1T2 + self.ep2) / 2
        self.constpT3 = (self.mohrsigma1T3 + self.ep3) / 2

        self.constqT1 = (self.mohrsigma1T1 - self.ep1) / 2
        self.constqT2 = (self.mohrsigma1T2 - self.ep2) / 2
        self.constqT3 = (self.mohrsigma1T3 - self.ep3) / 2

        self.pT = (self.constpT1, self.constpT2, self.constpT3)
        self.qT = (self.constqT1, self.constqT2, self.constqT3)

        self.linetotal = linregress(self.pT, self.qT)

        self.valphiT = degrees(asin(self.linetotal.slope))
        self.valcT = self.linetotal.intercept / cos(radians(self.valphiT))

    def mohrlineeffective(self):
        maxdev1 = max(self.deviatorstress1)
        imax1 = 0
        for imax1 in range(len(self.deviatorstress1)):
            imax1 += 1
            if maxdev1 == self.deviatorstress1[imax1]:
                break

        self.mohrsigma3E1 = self.sigma3E1[imax1]
        self.mohrsigma1E1 = self.sigma1E1[imax1]

        maxdev2 = max(self.deviatorstress2)
        imax2 = 0
        for imax2 in range(len(self.deviatorstress2)):
            imax2 += 1
            if maxdev2 == self.deviatorstress2[imax2]:
                break

        self.mohrsigma3E2 = self.sigma3E2[imax2]
        self.mohrsigma1E2 = self.sigma1E2[imax2]

        maxdev3 = max(self.deviatorstress3)
        imax3 = 0
        for imax3 in range(len(self.deviatorstress3)):
            imax3 += 1
            if maxdev3 == self.deviatorstress3[imax3]:
                break

        self.mohrsigma3E3 = self.sigma3E3[imax3]
        self.mohrsigma1E3 = self.sigma1E3[imax3]

        self.constaE1 = self.mohrsigma3E1 + ((self.mohrsigma1E1 - self.mohrsigma3E1) / 2)
        self.constaE2 = self.mohrsigma3E2 + ((self.mohrsigma1E2 - self.mohrsigma3E2) / 2)
        self.constaE3 = self.mohrsigma3E3 + ((self.mohrsigma1E3 - self.mohrsigma3E3) / 2)

        self.constrE1 = (self.mohrsigma1E1 - self.mohrsigma3E1) / 2
        self.constrE2 = (self.mohrsigma1E2 - self.mohrsigma3E2) / 2
        self.constrE3 = (self.mohrsigma1E3 - self.mohrsigma3E3) / 2

        self.constpE1 = (self.mohrsigma1E1 + self.mohrsigma3E1) / 2
        self.constpE2 = (self.mohrsigma1E2 + self.mohrsigma3E2) / 2
        self.constpE3 = (self.mohrsigma1E3 + self.mohrsigma3E3) / 2

        self.constqE1 = (self.mohrsigma1E1 - self.mohrsigma3E1) / 2
        self.constqE2 = (self.mohrsigma1E2 - self.mohrsigma3E2) / 2
        self.constqE3 = (self.mohrsigma1E3 - self.mohrsigma3E3) / 2

        self.pE = (self.constpE1, self.constpE2, self.constpE3)
        self.qE = (self.constqE1, self.constqE2, self.constqE3)

        self.lineeffective = linregress(self.pE, self.qE)

        self.valphiE = degrees(asin(self.lineeffective.slope))
        self.valcE = self.lineeffective.intercept / cos(radians(self.valphiE))

    def plotfigure(self):

        fig = plt.figure(constrained_layout=True, figsize=(8.3, 11.7), dpi=100)
        gs = fig.add_gridspec(4, 3)

        #add logo
        fig0 = fig.add_subplot(gs[0, 0])
        fig0.plot()
        fig0.set_xlim(0, 1)
        fig0.set_ylim(0, 1)
        fig0.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)
        fig0.tick_params(axis='y', which='both', left=False, right=False, labelleft=False)
        filelogo = mpimg.imread('itenas.png')
        logobox = OffsetImage(filelogo, zoom=0.4)
        logo = AnnotationBbox(logobox, (.5, .6), frameon=False)
        fig0.add_artist(logo)

        #add sample information
        infosample = AnchoredText('SAMPLE ID :' + ' ' + str(self.sampleid) + '\n'
                                  'DEPTH     :' + ' ' + str(round(self.depthfrom, 2)) + 
                                  ' ' + 
                                  '-' + 
                                  ' ' + 
                                  str(round(self.depthto, 2)) + 
                                  ' ' + 
                                  'm',
                                  prop=dict(size=12), frameon=False, loc='lower left')
        infosample.patch.set_boxstyle('round, pad=0., rounding_size=0.2')
        fig0.add_artist(infosample)


        #plot BValue
        fig1 = fig.add_subplot(gs[1, 0])        
        fig1.set(xlabel='Cell Pressure (kPa)', ylabel='B-Value', title='Saturation Build-Up')
        fig1.plot(self.sortcpvalue, self.bvalue, color='black', lw=1, marker='o', markersize=2) 
        fig1.grid(True, linestyle='-.')
        fig1.yaxis.set_major_locator(MultipleLocator(0.1))
        fig1.yaxis.set_minor_locator(AutoMinorLocator())
        fig1.minorticks_on()
        fig1.set_xlim(0, max(self.sortcpvalue) * 1.1)
        fig1.set_ylim(0, 1.0)

        #plot pwp vs strain
        fig2 = fig.add_subplot(gs[0, 1:])        
        fig2.set(xlabel='Strain (%)', ylabel='delta PWP (kPa)', title='delta PWP vs Strain')
        fig2.plot(self.strainvalue1, self.deltapwpshear1, color='black', lw=1, linestyle=':', label='Stage-1')
        fig2.plot(self.strainvalue2, self.deltapwpshear2, color='black', lw=1, linestyle='-.', label='Stage-2')
        fig2.plot(self.strainvalue3, self.deltapwpshear3, color='black', lw=1, linestyle='--', label='Stage-3')
        fig2.grid(True, linestyle='-.')
        fig2.xaxis.set_major_locator(MultipleLocator(1))
        fig2.xaxis.set_minor_locator(AutoMinorLocator())
        fig2.yaxis.set_minor_locator(AutoMinorLocator())
        fig2.set_xlim(0, max(self.strainvalue3) * 1.1)
        fig2.set_ylim(0, max(self.deltapwpshear3) * 1.1)
        fig2.legend(loc=2)

        #plot deviator stress vs strain
        fig3 = fig.add_subplot(gs[1, 1:])
        fig3.set(xlabel='Strain (%)', ylabel='Deviator Stress (kPa)', title='Deviator Stress vs Strain')
        fig3.plot(self.strainvalue1, self.deviatorstress1, color='black', lw=1, linestyle=':', label='Stage-1')
        fig3.plot(self.strainvalue2, self.deviatorstress2, color='black', lw=1, linestyle='-.', label='Stage-2')
        fig3.plot(self.strainvalue3, self.deviatorstress3, color='black', lw=1, linestyle='--', label='Stage-3')
        fig3.grid(True, linestyle='-.')
        fig3.xaxis.set_major_locator(MultipleLocator(1))
        fig3.xaxis.set_minor_locator(AutoMinorLocator())
        fig3.yaxis.set_minor_locator(AutoMinorLocator())
        fig3.set_xlim(0, max(self.strainvalue3) * 1.1)
        fig3.set_ylim(0, max(self.deviatorstress3) * 1.1)
        fig3.legend(loc=2)

        # plot Mohr's Diagram
        fig4 = fig.add_subplot(gs[2:, 0:])
        fig4.set(xlabel='Sigma (kPa)', ylabel='Tau (kPa)', title="Mohr's Diagram")
        fig4.grid(True, linestyle='-.')

        optionscala = [200, 300, 400, 500, 600, 700, 800, 900, 1000]
        maxgridX = optionscala[self.maxscale((self.mohrsigma1T3), optionscala) + 1]
        maxgridY = optionscala[self.maxscale((self.mohrsigma1T3), optionscala) + 1] * 60 / 100

        fig4.xaxis.set_major_locator(MultipleLocator(maxgridX / 10))
        fig4.yaxis.set_major_locator(MultipleLocator(maxgridX / 10))
        fig4.xaxis.set_minor_locator(AutoMinorLocator())
        fig4.yaxis.set_minor_locator(AutoMinorLocator())

        circle1T = plt.Circle((self.ep1 + (self.mohrsigma1T1 - self.ep1) / 2.0, 0), ((self.mohrsigma1T1 - self.ep1) / 2.0), color='black', linewidth=2 ,fill=False)
        circle2T = plt.Circle((self.ep2 + (self.mohrsigma1T2 - self.ep2) / 2.0, 0), ((self.mohrsigma1T2 - self.ep2) / 2.0), color='black', linewidth=2 ,fill=False)
        circle3T = plt.Circle((self.ep3 + (self.mohrsigma1T3 - self.ep3) / 2.0, 0), ((self.mohrsigma1T3 - self.ep3) / 2.0), color='black', linewidth=2 ,fill=False)
        
        circle1E = plt.Circle((self.mohrsigma3E1 + (self.mohrsigma1E1 - self.mohrsigma3E1) / 2.0, 0), ((self.mohrsigma1E1 - self.mohrsigma3E1) / 2.0), color='grey', linestyle='--', fill=False)
        circle2E = plt.Circle((self.mohrsigma3E2 + (self.mohrsigma1E2 - self.mohrsigma3E2) / 2.0, 0), ((self.mohrsigma1E2 - self.mohrsigma3E2) / 2.0), color='grey', linestyle='--', fill=False)
        circle3E = plt.Circle((self.mohrsigma3E3 + (self.mohrsigma1E3 - self.mohrsigma3E3) / 2.0, 0), ((self.mohrsigma1E3 - self.mohrsigma3E3) / 2.0), color='grey', linestyle='--', fill=False)

        fig4.add_artist(circle1T)
        fig4.add_artist(circle2T)
        fig4.add_artist(circle3T)

        fig4.add_artist(circle1E)
        fig4.add_artist(circle2E)
        fig4.add_artist(circle3E)

        xlinetotal = np.arange(0.0, self.mohrsigma1T3, 10)
        ylinetotal = self.linetotal.intercept + self.linetotal.slope * xlinetotal

        # xlineeffective = np.arange(0.0, self.mohrsigma1E3, 10)
        ylineeffective = self.lineeffective.intercept + self.lineeffective.slope * xlinetotal

        fig4.plot(xlinetotal, ylinetotal, lw=2, color='black', label='Total')
        fig4.plot(xlinetotal, ylineeffective, '--', lw=1, color='grey', label='Effective')

        

        fig4.set_xlim(0, maxgridX)
        fig4.set_ylim(0, maxgridY)

        infotext = AnchoredText('c =' + ' ' + str(round(self.valcT, 2)) + ' ' + 'kPa' + '\n'
                                'phi =' + ' ' + str(round(self.valphiT, 2)) + r'$^\circ$' + '\n'
                                "c' =" + ' ' + str(round(self.valcE, 2)) + ' ' + 'kPa' + '\n'
                                "phi' =" + ' ' + str(round(self.valphiE, 2)) + r'$^\circ$',
                                prop=dict(size=12), frameon=True, loc='upper left')
        infotext.patch.set_boxstyle('round, pad=0., rounding_size=0.2')
        fig4.add_artist(infotext)

        plt.legend(loc=1)
        fig.savefig(self.calcfilename.split('.')[0] + '.pdf', quality=100)

    def maxscale(self, valmax, optionscale):
        index = 0
        for i in range(len(optionscale)):
            if (optionscale[i] - valmax) <= 0:
                index += 1
            else:
                break
                index += 0
        return index

    def printoutput(self):

        ######List for Output in ListWidget######
        addlistoutput = ['Total Cohession Value : ',
                        'Total Phi Angle Value : ',
                        'Effective Cohession Value : ',
                        'Effective Phi Angle Value : ',
                        ]

        addlistvaloutput = [self.valcT,
                            self.valphiT,
                            self.valcE,
                            self.valphiE,
                            ]

        addlistunit = ['kPa',
                    'degree',
                    'kPa',
                    'degree',
                    ]
        
        headeroutput = ['Cell Pressure :',
                        'Back Pressure :',
                        'Effective Pressure :',
                        ]

        fillheaderoutput1 = [str(round(self.cp1)),
                            str(round(self.bp1)),
                            str(round(self.ep1)),
                            ]

        fillheaderoutput2 = [str(round(self.cp2)),
                            str(round(self.bp2)),
                            str(round(self.ep2)),
                            ]

        fillheaderoutput3 = [str(round(self.cp3)),
                            str(round(self.bp3)),
                            str(round(self.ep3)),
                            ]

        self.progressbar.setValue(0)
        for progi in range(0, 10000):
            self.progressbar.setValue(self.progressbar.value()+1)

        #####Format for setting Text style 1#####
        settingfontlistprintout1 = QFont()
        settingfontlistprintout1.setBold(True)
        settingfontlistprintout1.setWeight(75)

        #####Format for setting Text style 2#####
        settingfontlistprintout2 = QFont()
        settingfontlistprintout2.setBold(True)
        settingfontlistprintout2.setWeight(50)
        
        ######Header Printout######
        self.listprintout.addItem(
            'Licensed for: Institut Teknologi Nasional (ITENAS) by NGplot-(2020)'
        )

        self.listprintout1 = QListWidgetItem('Output :')
        self.listprintout1.setFont(settingfontlistprintout1)
        self.listprintout.addItem(self.listprintout1)        

        ######Display Summary Output Shear######
        for i in range(len(addlistoutput)):
            self.listprintout1 = QListWidgetItem(addlistoutput[i] + 
                                                str(round(addlistvaloutput[i], 2)) + 
                                                ' ' + 
                                                addlistunit[i])
            self.listprintout.addItem(self.listprintout1)

        
        self.listprintout.addItem(
            '---------------------------------------------------------------------------'
        )

        ######Displat List data from Saturation Stage######
        self.listprintout.addItem(
            '---SATURATION---'
        )

        self.listprintout.addItem(
            '---------------------------------------------------------------------------'
        )

        self.listprintout1 = QListWidgetItem('Cell Pressure (kPa) \t\t' + 'B-Value (-) \t')
        self.listprintout1.setFont(settingfontlistprintout2)
        self.listprintout.addItem(self.listprintout1)

        for i in range(len(self.bvalue)):
            self.listprintout1 = QListWidgetItem(str(round(self.sortcpvalue[i], 2)) +
                                                '\t\t\t' +
                                                str(round(self.bvalue[i], 2)) +
                                                '\t\t'
                                                )
            self.listprintout.addItem(self.listprintout1)

        ######Display List data from Shear Stage-1######   
        self.listprintout.addItem(
            '---------------------------------------------------------------------------'
        )

        self.listprintout.addItem(
            '---Shearing---'
        )

        self.listprintout.addItem(
            '---------------------------------------------------------------------------'
        )

        self.listprintout.addItem(
            'Stage 1 -'
        )

        for i in range(len(headeroutput)):
            self.listprintout.addItem(
                headeroutput[i] + ' ' + fillheaderoutput1[i] + ' ' + 'kPa'
            )
        
        
        self.listprintout.addItem(
            '---------------------------------------------------------------------------'
        )

        self.listprintout1 = QListWidgetItem('Strain (%) \t\t' + 'Deviator Stress (kPa) \t' + 'PWP (kPa)')
        self.listprintout1.setFont(settingfontlistprintout2)
        self.listprintout.addItem(self.listprintout1)

        for i in range(len(self.deviatorstress1)):
            self.listprintout1 = QListWidgetItem(str(round(self.strainvalue1[i], 2)) + 
                                                '\t\t' +
                                                str(round(self.deviatorstress1[i], 2)) +
                                                '\t\t\t' +
                                                str(round(self.deltapwpshear1[i], 2))
                                                )
            self.listprintout.addItem(self.listprintout1)
        
        ######Display List data from Shear Stage-2######   
        self.listprintout.addItem(
            '---------------------------------------------------------------------------'    
        ) 

        self.listprintout.addItem(
            'Stage 2 -'
        )

        for i in range(len(headeroutput)):
            self.listprintout.addItem(
                headeroutput[i] + ' ' + fillheaderoutput2[i] + ' ' + 'kPa'
            )

        self.listprintout.addItem(
            '---------------------------------------------------------------------------'    
        )

        self.listprintout1 = QListWidgetItem('Strain (%) \t\t' + 'Deviator Stress (kPa) \t' + 'PWP (kPa)')
        self.listprintout1.setFont(settingfontlistprintout2)
        self.listprintout.addItem(self.listprintout1)

        for i in range(len(self.deviatorstress2)):
            self.listprintout1 = QListWidgetItem(str(round(self.strainvalue2[i], 2)) +
                                                '\t\t' +
                                                str(round(self.deviatorstress2[i], 2)) +
                                                '\t\t\t' +
                                                str(round(self.deltapwpshear2[i], 2))
                                                )
            self.listprintout.addItem(self.listprintout1)

        
        ######Display List data from Shear Stage-3######   
        self.listprintout.addItem(
            '---------------------------------------------------------------------------'    
        )

        self.listprintout.addItem(
            'Stage 3 -'
        )

        for i in range(len(headeroutput)):
            self.listprintout.addItem(
                headeroutput[i] + ' ' + fillheaderoutput3[i] + ' ' + 'kPa'
            )
        
        self.listprintout.addItem(
            '---------------------------------------------------------------------------'    
        )

        self.listprintout1 = QListWidgetItem('Strain (%) \t\t' + 'Deviator Stress (kPa) \t' + 'PWP (kPa)')
        self.listprintout1.setFont(settingfontlistprintout2)
        self.listprintout.addItem(self.listprintout1)

        for i in range(len(self.deviatorstress3)):
            self.listprintout1 = QListWidgetItem(str(round(self.strainvalue3[i], 2)) +
                                                '\t\t' +
                                                str(round(self.deviatorstress3[i], 2)) +
                                                '\t\t\t' +
                                                str(round(self.deltapwpshear3[i], 2))
                                                )
            self.listprintout.addItem(self.listprintout1)

        self.listprintout.addItem(
            '-----by NGplot-----'
        )       

    def printrawdata(self):
        
        pdf = AddHeader()

        #create the special value {nb}
        pdf.alias_nb_pages()
        
        pdf.add_page()
        
        pdf.set_font('Courier', 'BU', size=11)
        pdf.cell(50, 5, txt='Output:', border=0, ln=1, align='L')
        
        pdf.cell(50, 5, txt='', border=0, ln=1, align='L')

        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Saturation', border=0, ln=1, align='L')

        pdf.cell(50, 5, txt='Cell Pressure (kPa)', border=0, ln=0, align='L')
        pdf.cell(50, 5, txt='B-Value (-)', border=0, ln=1, align='L')
        pdf.set_font('Courier', size=10)
        for i in range(len(self.bvalue)):
            pdf.cell(50, 5, txt=str(round(self.sortcpvalue[i])), border=0, ln=0, align='L')
            pdf.cell(50, 5, txt=str(round(self.bvalue[i], 2)), border=0, ln=1, align='L')

        pdf.cell(50, 5, txt='', border=0, ln=1, align='L')
        
        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Shearing', border=0, ln=1, align='L')

        pdf.cell(50, 5, txt='Stage - 1', border=0, ln=1, align='L')

        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Strain (%)', ln=0, align='L')
        pdf.cell(50, 5, txt='Deviator Stress (kPa)', ln=0, align='L')
        pdf.cell(50, 5, txt='PWP (kPa)', ln=1, align='L')
        pdf.set_font('Courier', size=10)
        for i in range(len(self.deviatorstress1)):
            pdf.cell(50, 5, txt=str(str(round(self.strainvalue1[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.deviatorstress1[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.deltapwpshear1[i], 2))), ln=1, align='L')

        pdf.cell(50, 5, txt='', border=0, ln=1, align='L')
        pdf.cell(50, 5, txt='Stage - 2', border=0, ln=1, align='L')

        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Strain (%)', ln=0, align='L')
        pdf.cell(50, 5, txt='Deviator Stress (kPa)', ln=0, align='L')
        pdf.cell(50, 5, txt='PWP (kPa)', ln=1, align='L')
        pdf.set_font('Courier', size=10)
        for i in range(len(self.deviatorstress2)):
            pdf.cell(50, 5, txt=str(str(round(self.strainvalue2[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.deviatorstress2[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.deltapwpshear2[i], 2))), ln=1, align='L')


        pdf.cell(50, 5, txt='', border=0, ln=1, align='L')
        pdf.cell(50, 5, txt='Stage - 3', border=0, ln=1, align='L')

        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Strain (%)', ln=0, align='L')
        pdf.cell(50, 5, txt='Deviator Stress (kPa)', ln=0, align='L')
        pdf.cell(50, 5, txt='PWP (kPa)', ln=1, align='L')
        pdf.set_font('Courier', size=10)
        for i in range(len(self.deviatorstress3)):
            pdf.cell(50, 5, txt=str(str(round(self.strainvalue3[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.deviatorstress3[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.deltapwpshear3[i], 2))), ln=1, align='L')


        pdf.cell(50, 5, txt='', border=0, ln=1, align='L')
        
        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Consolidation', border=0, ln=1, align='L')

        pdf.cell(50, 5, txt='Stage - 1', border=0, ln=1, align='L')

        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Time (min)', ln=0, align='L')
        pdf.cell(50, 5, txt='PWP (kPa)', ln=0, align='L')
        pdf.cell(50, 5, txt='Vol Channge (cm3)', ln=1, align='L')
        pdf.set_font('Courier', size=10)
        for i in range(len(self.time1)):
            pdf.cell(50, 5, txt=str(str(round(self.time1[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.pwpconsol1[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.volchangeconsol1[i], 2))), ln=1, align='L')

        pdf.cell(50, 5, txt='', border=0, ln=1, align='L')

        pdf.cell(50, 5, txt='Stage - 2', border=0, ln=1, align='L')

        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Time (min)', ln=0, align='L')
        pdf.cell(50, 5, txt='PWP (kPa)', ln=0, align='L')
        pdf.cell(50, 5, txt='Vol Channge (cm3)', ln=1, align='L')
        pdf.set_font('Courier', size=10)
        for i in range(len(self.time2)):
            pdf.cell(50, 5, txt=str(str(round(self.time2[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.pwpconsol2[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.volchangeconsol2[i], 2))), ln=1, align='L')


        pdf.cell(50, 5, txt='', border=0, ln=1, align='L')

        pdf.cell(50, 5, txt='Stage - 3', border=0, ln=1, align='L')

        pdf.set_font('Courier', 'B', size=10)
        pdf.cell(50, 5, txt='Time (min)', ln=0, align='L')
        pdf.cell(50, 5, txt='PWP (kPa)', ln=0, align='L')
        pdf.cell(50, 5, txt='Vol Channge (cm3)', ln=1, align='L')
        pdf.set_font('Courier', size=10)
        for i in range(len(self.time3)):
            pdf.cell(50, 5, txt=str(str(round(self.time3[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.pwpconsol3[i], 2))), ln=0, align='L')
            pdf.cell(50, 5, txt=str(str(round(self.volchangeconsol3[i], 2))), ln=1, align='L')

        pdf.cell(50, 10, txt='', border=0, ln=1, align='L')
        pdf.cell(190, 5, txt='-----by NGplot-----', ln=1, align='C')


        self.pdfname = ((self.openfiletext.text()).split('.'))[0] + str('temp.pdf')

        pdf.output(self.pdfname)
        pdf.close()

    def mergerawfig(self):

        self.readfile = PdfFileReader(self.calcfilename.split('.')[0] + '.pdf')
        self.mergefile = PdfFileMerger()
        
        self.mergefile.append(self.readfile)
        self.mergefile.append(self.pdfname)

        with open(self.calcfilename.split('.')[0] + '.pdf', 'wb') as fileobj:
            self.mergefile.write(fileobj)
        
        self.mergefile.close()

        self.deletefigure()

    def showresultbuttonClick(self):
        os.startfile(self.calcfilename.split('.')[0] + '.pdf')
        
    def deletefigure(self):
        os.remove(self.pdfname)

    def setInputCenter(self):
        desktop = QDesktopWidget()
        screenwidth = desktop.screen().width()
        screenheight = desktop.screen().height()

        self.setGeometry((screenwidth - self.width()) // 2,
                         (screenheight - self.height()) // 2,
                         (self.width()),
                         (self.height()))

#class for Header and Footer
class AddHeader(FPDF):
    '''add header and footer'''
    def header(self):
        # #set up logo
        self.image('itenas.png', 10, 5, 30)
        self.set_font('Courier', 'BU', size=20)
        self.cell(190, 10, txt='INSTITUT TEKNOLOGI NASIONAL', border=0, ln=1, align='C')
        
        self.cell(190, 5, txt='', border=0, ln=1, align='C')

        self.set_font('Courier', 'B', size=14)
        self.cell(190, 5, txt='Triaxial Consolidated Undrained', border=0, ln=1, align='C')
        self.cell(190, 5, txt=32*'-', border=0, ln=1, align='C')
        self.cell(190, 5, txt='', border=0, ln=1, align='C')

        self.set_font('Courier', size=12)
        self.cell(190, 5, txt='', border=0, ln=1, align='C')
        self.cell(190, 5, txt='SAMPLE ID  :' + ' ' + str(idsample), border=0, ln=1, align='C')
        self.cell(190, 5, txt='\t\t\t\t' + 'DEPTH      :' + ' ' + str(fromdepth) + ' ' + '-' + ' ' + str(todepth) + ' ' + 'm', border=0, ln=1, align='C')

        # #line break
        # self.ln(20)
            ###
        # Set up a logo
        # self.image('itenas.png', 10, 8, 33)
        # self.set_font('Arial', 'B', 15)
        
        # Add an address
        # self.cell(100)
        # self.cell(0, 5, 'Mike Driscoll', ln=1)
        # self.cell(100)
        # self.cell(0, 5, '123 American Way', ln=1)
        # self.cell(100)
        # self.cell(0, 5, 'Any Town, USA', ln=1)
        
        # Line break
        self.ln(10)

    def footer(self):
        self.set_y(-10)
        
        self.set_font('Courier', 'I', size=8)

        # add a file address
        addressfile = str(pathfile)
        self.cell(0, 10, addressfile, border=0, ln=0, align='L')

        #add a page number
        page = str(self.page_no() + 1)      #start from 2
        self.cell(0, 10, page, border=0, ln=0, align='R')

        # self.set_y(-10)        
        # self.set_font('Arial', 'I', 8)
        # # Add a page number
        # page = 'Page ' + str(self.page_no()) + '/{nb}'
        # self.cell(0, 10, page, 0, 0, 'R')

if __name__ =='__main__':
    a = QApplication(sys.argv)

    form = MainForm()
    form.show()
    a.exec_()
